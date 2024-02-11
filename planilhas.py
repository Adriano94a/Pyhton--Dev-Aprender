#Manipulando planilhas
#Sheets - São as páginas de uma planilha
#Workboot - é o arquivo que contém estas páginas
import openpyxl
# Criar planilha
Plan_test = openpyxl.Workbook()
Plan_test.create_sheet('Frutas')
Plan_test.create_sheet('Legumes')
Plan_test.create_sheet('Sementes')
print(Plan_test.sheetnames)
cabecalho_frutas = ['Titulo', 'Localização', 'Preço']
sheet_frutas = Plan_test.get_sheet_by_name('Frutas')
sheet_frutas.append(cabecalho_frutas)

frutas = [['Uva', 'Mercado', 5],
          ['Melão', 'Mercado', 15],
          ['Amora', 'Mercado', 7]]

for fruta in frutas:
    sheet_frutas.append(fruta)

Plan_test.save('Dados Supermercado.xlsx')
